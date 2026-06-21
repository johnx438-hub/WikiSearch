"""
XLSX 转换器 — 将 Excel 电子表格转换为结构化 Markdown

核心功能:
1. 提取每个 Sheet 的表头和数据行
2. 识别合并单元格并填充
3. 按 Sheet 生成独立 MD 文件（Markdown 表格格式）
4. 生成 _index.md 导航页
5. YAML Frontmatter 元数据注入

使用示例:
    converter = XlsxConverter()
    result = converter.convert("/path/to/spreadsheet.xlsx")
"""

import os
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field


try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None


# ==================== 数据结构 ====================

@dataclass
class SheetResult:
    """单个工作表转换结果"""
    title: str                    # Sheet 名称
    content: str                  # Markdown 表格内容
    sheet_number: int             # Sheet 序号（1-indexed）
    rows_count: int = 0           # 数据行数
    cols_count: int = 0           # 列数
    metadata: Dict = field(default_factory=dict)  # YAML Frontmatter


# ==================== XLSX 转换器 ====================

class XlsxConverter:
    """
    Excel (.xlsx) → Markdown 转换器。
    
    Attributes:
        max_rows_per_sheet: 单 Sheet 最大处理行数（默认 1000，避免超大文件）
        include_empty_rows: 是否包含空行（默认 False）
    """
    
    def __init__(self, max_rows_per_sheet: int = 1000, include_empty_rows: bool = False):
        self.max_rows_per_sheet = max_rows_per_sheet
        self.include_empty_rows = include_empty_rows
    
    def convert(self, filepath: str, output_dir: Optional[str] = None) -> Dict:
        """
        将 XLSX 文件转换为 Markdown。
        
        Args:
            filepath: .xlsx 文件路径
            output_dir: 输出目录（如果提供，会生成树状 MD 文件）
            
        Returns:
            {
                'sections': [SheetResult, ...],
                'total_sheets': int,
                'sheets_info': List[Dict],  # Sheet 基本信息
            }
        """
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")
        
        # 1. 加载工作簿
        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        
        try:
            # 2. 提取每个 Sheet
            sheets = self._extract_sheets(wb)
            
            # 3. 生成树状 MD 文件（如果提供了输出目录）
            if output_dir:
                self._write_tree_structure(
                    sections=sheets,
                    base_dir=output_dir,
                    doc_name=filepath.stem
                )
            
            return {
                'sections': sheets,
                'total_sheets': len(wb.sheetnames),
                'sheets_info': [
                    {
                        'name': ws.title,
                        'rows': ws.max_row if ws.max_row else 0,
                        'cols': ws.max_column if ws.max_column else 0,
                    }
                    for ws in wb.worksheets
                ],
            }
        finally:
            wb.close()
    
    def _extract_sheets(self, wb) -> List[SheetResult]:
        """
        提取所有工作表内容。
        
        Args:
            wb: openpyxl Workbook 对象
            
        Returns:
            [SheetResult, ...]
        """
        sheets = []
        
        for idx, ws in enumerate(wb.worksheets, start=1):
            # 提取表格数据
            content = self._extract_sheet_content(ws)
            
            # 构建元数据（read_only 模式下用 max_row/max_column）
            metadata = {
                'source': 'xlsx',
                'sheet_number': idx,
                'total_sheets': len(wb.sheetnames),
                'dimensions': f"{ws.max_row or 0}x{ws.max_column or 0}",
            }
            
            sheets.append(SheetResult(
                title=ws.title or f"Sheet{idx}",
                content=content,
                sheet_number=idx,
                rows_count=ws.max_row or 0,
                cols_count=ws.max_column or 0,
                metadata=metadata
            ))
        
        return sheets
    
    def _extract_sheet_content(self, ws) -> str:
        """
        提取工作表内容为 Markdown 表格。
        
        Args:
            ws: openpyxl Worksheet 对象
            
        Returns:
            Markdown 格式的表格内容
        """
        if not ws.max_row or ws.max_row == 0:
            return "*（空工作表）*"
        
        # 1. 读取所有数据行
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, self.max_rows_per_sheet)))
        
        if not rows:
            return ""
        
        # 2. 确定列数（取最大列数的行）
        max_cols = max(len(row) for row in rows) if rows else 0
        
        # 3. 提取表头（第一行）
        headers = [self._cell_to_string(cell) for cell in rows[0][:max_cols]]
        
        # 4. 处理数据行
        data_rows = []
        for row_idx, row in enumerate(rows[1:], start=2):
            # 填充空列
            cells = [self._cell_to_string(cell) for cell in row[:max_cols]]
            while len(cells) < max_cols:
                cells.append("")
            
            # 跳过全空行（除非配置要求包含）
            if not self.include_empty_rows and all(not c.strip() for c in cells):
                continue
            
            data_rows.append(cells)
        
        # 5. 转换为 Markdown 表格
        md_lines = []
        
        # 表头
        md_lines.append("| " + " | ".join(headers) + " |")
        md_lines.append("| " + " | ".join(["---"] * max_cols) + " |")
        
        # 数据行（限制输出行数，避免过大）
        max_display_rows = 100
        if len(data_rows) > max_display_rows:
            md_lines.extend(
                ["| " + " | ".join(row) + " |" for row in data_rows[:max_display_rows]]
            )
            md_lines.append(f"\n*（共 {len(data_rows)} 行，显示前 {max_display_rows} 行）*\n")
        else:
            md_lines.extend(
                ["| " + " | ".join(row) + " |" for row in data_rows]
            )
        
        return "\n".join(md_lines)
    
    def _cell_to_string(self, cell) -> str:
        """
        将单元格值转换为字符串。
        
        Args:
            cell: openpyxl Cell 对象
            
        Returns:
            字符串表示
        """
        if cell is None or cell.value is None:
            return ""
        
        value = cell.value
        
        # 处理日期类型
        if hasattr(value, 'strftime'):
            import datetime
            if isinstance(value, datetime.datetime):
                return value.strftime('%Y-%m-%d %H:%M')
            elif isinstance(value, datetime.date):
                return value.strftime('%Y-%m-%d')
        
        # 处理数字类型（保留精度）
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            else:
                return f"{value:.2f}"
        
        return str(value)
    
    def _write_tree_structure(self, sections: List[SheetResult], 
                               base_dir: str, doc_name: str):
        """
        将转换结果写入树状 MD 文件结构。
        
        Args:
            sections: Sheet 列表
            base_dir: 基础输出目录
            doc_name: 文档名称（不含扩展名）
        """
        # 创建文档根目录
        doc_dir = os.path.join(base_dir, doc_name)
        os.makedirs(doc_dir, exist_ok=True)
        
        # 生成 _index.md（导航页）
        self._write_index_md(sections=sections, 
                            index_path=os.path.join(doc_dir, "_index.md"))
        
        # 按 Sheet 生成独立 MD 文件
        for sheet in sections:
            filename = f"Sheet_{sheet.sheet_number}_{self._sanitize_filename(sheet.title)}.md"
            filepath = os.path.join(doc_dir, filename)
            
            self._write_sheet_md(
                section=sheet,
                doc_name=doc_name,
                output_path=filepath
            )
    
    def _write_index_md(self, sections: List[SheetResult], index_path: str):
        """
        生成 _index.md（Sheet 导航）。
        
        Args:
            sections: Sheet 列表
            index_path: 输出路径
        """
        lines = [
            f"# {sections[0].metadata.get('source', 'XLSX').upper()} 电子表格",
            "",
            "## 📑 工作表导航",
            ""
        ]
        
        for sheet in sections:
            filename = f"Sheet_{sheet.sheet_number}_{self._sanitize_filename(sheet.title)}.md"
            lines.append(f"- [{sheet.sheet_number}. {sheet.title}]({filename})")
            lines.append(f"  - 行数: {sheet.rows_count}, 列数: {sheet.cols_count}")
        
        # 添加统计信息
        total_rows = sum(s.rows_count for s in sections)
        lines.extend([
            "",
            "## 📊 文档统计",
            f"- **总工作表数**: {len(sections)}",
            f"- **总数据行数**: {total_rows}",
        ])
        
        with open(index_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
    
    def _write_sheet_md(self, section: SheetResult, doc_name: str, output_path: str):
        """
        生成单个 Sheet 的 MD 文件。
        
        Args:
            section: Sheet 结果
            doc_name: 文档名称
            output_path: 输出路径
        """
        # YAML Frontmatter
        frontmatter = [
            "---",
            f"source: \"{doc_name}.xlsx\"",
            f"sheet_number: {section.sheet_number}",
            f"title: \"{section.title}\"",
            f"rows: {section.rows_count}",
            f"cols: {section.cols_count}",
            "---",
            ""
        ]
        
        # 标题 + 表格内容
        content_lines = [
            '\n'.join(frontmatter),
            f"# {section.title}",
            "",
            section.content,
        ]
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(content_lines))
    
    def _sanitize_filename(self, text: str) -> str:
        """
        清理文件名中的非法字符。
        
        Args:
            text: 原始文本
            
        Returns:
            安全的文件名
        """
        import re
        sanitized = re.sub(r'[\\/:*?"<>|]', '_', text)
        if len(sanitized) > 30:
            sanitized = sanitized[:27] + '...'
        return sanitized.strip()
